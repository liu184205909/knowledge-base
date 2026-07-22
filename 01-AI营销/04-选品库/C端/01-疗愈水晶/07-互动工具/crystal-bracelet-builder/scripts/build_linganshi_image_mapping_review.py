from __future__ import annotations

import hashlib
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image


ROOT = Path(r"D:\Code\knowledge-base\01-AI营销\04-选品库\C端\01-疗愈水晶\07-互动工具\crystal-bracelet-builder")
RESEARCH = ROOT / "data" / "v3" / "research"
CATALOG = RESEARCH / "linganshi-material-catalog-20260715.xlsx"
IMAGE_SOURCE = Path(r"C:\Users\Dylan\Desktop\灵感石验室")
OUTPUT = RESEARCH / "linganshi-image-mapping-review-20260720.xlsx"


def style_sheet(sheet, table_name: str) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if sheet.max_row > 1:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in sheet.columns:
        letter = column[0].column_letter
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[letter].width = min(max(width, 10), 48)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_rows(sheet, headers: list[str], rows: list[dict[str, object]], table_name: str) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    style_sheet(sheet, table_name)


def read_catalog() -> tuple[dict[str, dict[str, object]], list[str], list[dict[str, object]]]:
    workbook = load_workbook(CATALOG, read_only=True, data_only=True)
    sheet = workbook["变体明细"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(rows)]
    index = {header: position for position, header in enumerate(headers)}
    products: dict[str, dict[str, object]] = {}
    variant_rows: list[dict[str, object]] = []
    for row in rows:
        variant_row = {header: row[position] for header, position in index.items()}
        variant_rows.append(variant_row)
        name = str(row[index["商品名称"]] or "").strip()
        if not name:
            continue
        product = products.setdefault(name, {
            "一级类目": str(row[index["一级类目"]] or ""),
            "二级类目": str(row[index["二级类目"]] or ""),
            "规格": set(),
            "价格": set(),
        })
        size = str(row[index["尺寸_mm"]] or "").strip()
        price = str(row[index["单价_CNY"]] or "").strip()
        if size:
            product["规格"].add(size)
        if price:
            product["价格"].add(price)
    return products, headers, variant_rows


def read_image(path: Path) -> dict[str, object]:
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    with Image.open(path) as image:
        detected_format = str(image.format or "unknown").lower()
        has_alpha = image.mode in {"RGBA", "LA"} or "transparency" in image.info
        width, height = image.size
    extension = path.suffix.lower().lstrip(".")
    return {
        "图片文件名": path.name,
        "图片绝对路径": str(path),
        "SHA256": digest,
        "宽_px": width,
        "高_px": height,
        "透明通道": "yes" if has_alpha else "no",
        "文件扩展名": extension,
        "真实图像格式": detected_format,
        "扩展名一致": "yes" if extension == detected_format else "no",
    }


def main() -> None:
    if not CATALOG.is_file():
        raise FileNotFoundError(CATALOG)
    if not IMAGE_SOURCE.is_dir():
        raise FileNotFoundError(IMAGE_SOURCE)
    if ROOT not in OUTPUT.parents:
        raise RuntimeError(f"Output escaped project root: {OUTPUT}")

    products, variant_headers, variant_rows = read_catalog()
    image_paths = sorted(
        (path for path in IMAGE_SOURCE.iterdir() if path.is_file() and path.suffix.lower() in {".png", ".webp"}),
        key=lambda item: item.name,
    )
    images = [read_image(path) for path in image_paths]
    hash_groups: dict[str, list[str]] = defaultdict(list)
    for image in images:
        hash_groups[str(image["SHA256"])].append(str(image["图片文件名"]))

    product_rows: list[dict[str, object]] = []
    image_product_names: set[str] = set()
    for image in images:
        image_name = Path(str(image["图片文件名"])).stem
        image_product_names.add(image_name)
        product = products.get(image_name)
        shared_files = hash_groups[str(image["SHA256"])]
        material_key = "linganshi-" + hashlib.sha1(image_name.encode("utf-8")).hexdigest()[:12]
        variant_count = len(product["规格"]) if product else 0
        product_rows.append({
            "研究商品键": material_key,
            "商品名称": image_name,
            **image,
            "一级类目": product["一级类目"] if product else "",
            "二级类目": product["二级类目"] if product else "",
            "已核验尺寸": " | ".join(sorted(product["规格"])) if product else "",
            "已记录价格_CNY": " | ".join(sorted(product["价格"])) if product else "",
            "已核验Variant数": variant_count,
            "商品建档状态": "product-mapped",
            "Variant覆盖状态": "existing-variants" if product else "pending-collection",
            "图片匹配方式": "user-confirmed-filename",
            "共用图片状态": "shared-image-intentional" if len(shared_files) > 1 else "unique-image",
            "共用图片文件": " | ".join(shared_files) if len(shared_files) > 1 else "",
            "本地开发素材状态": "temporary-draft-enabled",
            "生产导入状态": "pending-variant-and-rule-review",
            "说明": "商品名称已按文件名确认；缺少的 Variant、分类、方向、适配珠径与占位需从小程序补采",
        })

    pending_variant_rows = [
        {
            "研究商品键": row["研究商品键"],
            "商品名称": row["商品名称"],
            "图片文件名": row["图片文件名"],
            "待补字段": "一级类目 | 二级类目 | Variant尺寸 | 价格 | 占位长度 | 重量 | 方向 | 适配珠径",
            "补采状态": "pending-mini-program-review",
        }
        for row in product_rows if row["Variant覆盖状态"] == "pending-collection"
    ]

    legacy_unmatched_rows = []
    for name in sorted(set(products) - image_product_names):
        product = products[name]
        legacy_unmatched_rows.append({
            "商品名称": name,
            "一级类目": product["一级类目"],
            "二级类目": product["二级类目"],
            "已核验尺寸": " | ".join(sorted(product["规格"])),
            "已记录价格_CNY": " | ".join(sorted(product["价格"])),
            "处理状态": "alias-or-missing-image-review",
        })

    existing_variant_products = sum(1 for row in product_rows if row["Variant覆盖状态"] == "existing-variants")
    shared_image_rows = sum(1 for row in product_rows if row["共用图片状态"] == "shared-image-intentional")
    wrong_extension = sum(1 for row in product_rows if row["扩展名一致"] == "no")
    summary_rows = [
        {"指标": "原工作簿 Variant 行", "数量": len(variant_rows), "状态": "完整保留到已有Variant页"},
        {"指标": "图片商品主记录", "数量": len(product_rows), "状态": "235/235 已按文件名建档"},
        {"指标": "已有 Variant 的图片商品", "数量": existing_variant_products, "状态": "保留原尺寸和价格"},
        {"指标": "待补 Variant 的图片商品", "数量": len(pending_variant_rows), "状态": "需从小程序继续补采"},
        {"指标": "旧表未对应图片商品", "数量": len(legacy_unmatched_rows), "状态": "别名或缺图待复核"},
        {"指标": "确认共用图片的商品行", "数量": shared_image_rows, "状态": "白水晶与高品净体白水晶保留为独立商品"},
        {"指标": "扩展名与真实格式不一致", "数量": wrong_extension, "状态": "应为 0"},
        {"指标": "本地开发可用图片商品", "数量": len(product_rows), "状态": "temporary-draft-enabled"},
    ]

    workbook = Workbook()
    summary = workbook.active
    summary.title = "映射摘要"
    write_rows(summary, ["指标", "数量", "状态"], summary_rows, "MappingSummary")
    write_rows(workbook.create_sheet("商品主表"), [
        "研究商品键", "商品名称", "图片文件名", "图片绝对路径", "SHA256", "宽_px", "高_px", "透明通道",
        "文件扩展名", "真实图像格式", "扩展名一致", "一级类目", "二级类目", "已核验尺寸",
        "已记录价格_CNY", "已核验Variant数", "商品建档状态", "Variant覆盖状态", "图片匹配方式",
        "共用图片状态", "共用图片文件", "本地开发素材状态", "生产导入状态", "说明",
    ], product_rows, "ProductMaster")
    variant_output_rows = []
    for row in variant_rows:
        name = str(row.get("商品名称") or "").strip()
        output_row = dict(row)
        output_row["图片商品对应状态"] = "exact-name" if name in image_product_names else "alias-or-missing-image-review"
        variant_output_rows.append(output_row)
    write_rows(workbook.create_sheet("已有Variant"), variant_headers + ["图片商品对应状态"],
               variant_output_rows, "ExistingVariants")
    write_rows(workbook.create_sheet("图片商品待补Variant"), [
        "研究商品键", "商品名称", "图片文件名", "待补字段", "补采状态",
    ], pending_variant_rows, "PendingProductVariants")
    write_rows(workbook.create_sheet("旧表未匹配商品"), [
        "商品名称", "一级类目", "二级类目", "已核验尺寸", "已记录价格_CNY", "处理状态",
    ], legacy_unmatched_rows, "LegacyUnmatchedProducts")
    workbook.save(OUTPUT)

    verified = load_workbook(OUTPUT, read_only=True, data_only=True)
    print(OUTPUT)
    print(" | ".join(f"{sheet.title}:{sheet.max_row - 1}" for sheet in verified.worksheets))
    print(f"products={len(product_rows)} variants={len(variant_rows)} "
          f"existing_variant_products={existing_variant_products} pending_variant_products={len(pending_variant_rows)} "
          f"legacy_unmatched_products={len(legacy_unmatched_rows)} shared_image_rows={shared_image_rows} "
          f"wrong_extension={wrong_extension} local_draft_enabled={len(product_rows)}")


if __name__ == "__main__":
    main()
