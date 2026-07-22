"""Write directly observed Linganshi mini-program card values into the one master workbook.

This is deliberately a narrow reconciliation pass: it only writes values actually
observed in the mini-program UI on 2026-07-21.  A card that visibly has a size
switch is retained as a partial observation until both endpoints are tested.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import re

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / "data" / "v3" / "research" / "linganshi-image-mapping-review-20260720.xlsx"
IMAGES = Path(r"C:\Users\Dylan\Desktop\灵感石验室")
OBSERVED_ON = "2026-07-21"
EVIDENCE = f"A-mini-program-ui-{OBSERVED_ON}"

# Wording-only bridges from a visible card to one manually-confirmed image file.
# They are explicit rather than fuzzy; material substitutions never enter here.
APP_NAME_TO_IMAGE_NAME = {
    "五叶饶戒": "五叶绕戒",
    "四叶草（可转动）": "四叶草",
    "四环绕戒隔环": "四环绕戒隔环金",
    "浅色透体红胶花": "浅色透底红胶花",
    "浅色透体黄胶花": "浅色透底黄胶花",
    "白水晶方糖切面": "白方晶方糖切面",
    "白水晶弧形切面珠": "白水晶弧形切面",
    "白水晶椭圆切面珠": "白水晶椭圆切面",
    "芬达色胶花": "芬达色浇花",
    "高品黄塔晶随型 中": "高品黄塔晶随型",
    "锆石长隔环金": "锆石长隔珠金",
}

# These two local products were actively checked in the current app rather than
# silently left as a missing row.  They have no safe price/size substitution.
UNRESOLVED_IMAGE_AUDIT = {
    "琥珀晶随型 小": "not found when searched in the current 随型 tab",
    "白曜小隔片": "current 黑曜小隔片 card is visually and materially different",
}


def norm(value: str) -> str:
    return re.sub(r"[\s\-_/（）()·,.，。]+", "", str(value or "").strip()).lower()


def rows(raw: str, source_tab: str, component: str, switchable: bool) -> list[dict[str, object]]:
    result: list[dict[str, object]] = []
    for line in raw.strip().splitlines():
        parts = [part.strip() for part in line.split("|")]
        name, size, price = parts[:3]
        result.append({
            "name": name,
            "size": float(size),
            "price": float(price),
            "source_tab": source_tab,
            "component": component,
            "switchable": switchable,
            "complete": not switchable,
            "fit": parts[3] if len(parts) > 3 else "",
        })
    return result


# Current cards read from the running Linganshi mini-program.  Every Crystal card
# below showed a size +/- control in the captured list; its displayed default is
# useful evidence, but not a claim that it is the only Variant.
CRYSTALS = rows(
    """
白阿塞|10|5
高品净体白水晶|10|11
白水晶|10|7
莫粉|10|17
西柚粉晶|10|8
粉晶|10|13
紫马粉|10|17
薰衣草|10|12
巴西紫水晶|10|25
高品玻利维亚紫水晶|10|45
乌拉圭紫水晶|10|25
玻利维亚紫水晶|10|20
魔鬼蓝海蓝宝|10|81
冰川蓝海蓝宝|10|30
蓝天白云海蓝宝|10|16
柠檬黄水晶|10|15
高透黄水晶|10|25
黄塔晶|10|12
绿水晶|9|8
浅色透体黄胶花|10|25
浅色透体红胶花|12|25
牛血红胶花|10|19
浅色红胶花|12|25
黄胶花|10|14
芬达色胶花|10|14
白幽灵|10|17
满天星绿幽灵|10|17
千层彩幽灵|12|31
奶油体雪花幽灵|10|11
半盆绿幽灵|10|55
透体雪花幽灵|10|31
千层绿幽灵|10|32
浅色草莓晶|10|7
草莓晶|10|9
蓝晶石|10|23
深茶晶|10|10
浅茶晶|10|10
透体黑金超|10|27
黑金超七色浓|10|20
满发黑金超七|10|13
高品金发晶|8|35
黑发晶浅色|10|23
透体黑发晶|10|75
金发晶|10|28
满发绿发晶|10|12
黑发晶深色|10|23
浅色绿发晶|10|24
闪灵|12|30
""",
    "水晶",
    "crystal",
    True,
)

# The first card was fully stepped down/up in the UI.  These are the only three
# size/price pairs in this script asserted as boundary-complete for that material.
CRYSTALS.extend([
    {"name": "白阿塞", "size": 8.0, "price": 4.0, "source_tab": "水晶", "component": "crystal", "switchable": True, "complete": True},
    {"name": "白阿塞", "size": 10.0, "price": 5.0, "source_tab": "水晶", "component": "crystal", "switchable": True, "complete": True},
    {"name": "白阿塞", "size": 12.0, "price": 10.0, "source_tab": "水晶", "component": "crystal", "switchable": True, "complete": True},
])

NATURAL_SWITCHABLE = rows(
    """
岫玉|10|9
青提奶盖|10|10
紫锂辉|10|18
蓝虎眼|10|14
黄虎眼|10|13
黄萤石|10|12
太阳石|10|20
月光石|10|22
金太阳|8|27
透体灰月光|10|30
灰月光|10|28
双眼金曜石|10|6
双眼银曜石|10|6
高透冰曜石|8|4
南红玛瑙|10|7
绿龙晶|10|16
紫云母|10|20
朱砂|10|8
葡萄石|10|21
""",
    "天然石",
    "crystal",
    True,
)
NATURAL_FIXED = rows(
    """
天河石|9|7
绿英石|10|4
""",
    "天然石",
    "crystal",
    False,
)

# 随型属于 Accessories，而非 Crystals；已显示 +/- 的卡仍需要做边界扫描。
SHAPED_FIXED = rows(
    """
高品黄塔晶随型 中|14|42
白水晶方糖切面|11|14
粉水晶随型|15|18
柠檬晶菠萝珠|12|24
蜜蜡桶珠|6|10
银曜石平安扣|8|13
琥珀隔珠|10|19
粉水晶平安扣|6|14
白水晶平安扣|8|18
白水晶切面隔珠|5|10
茶晶切面隔珠|7|15
黄水晶切面隔珠|7|15
黄虎眼隔珠|4|6
青提奶盖平安扣|12|13
黄塔晶随型 小|10|23
薰衣草随型|17|27
紫水晶随型|15|28
黄塔晶随型 大|16|16
白水晶随型|18|12
绿幽灵方糖|12|21
黄胶花方糖|12|21
黑发晶方糖|12|52
银曜石方糖|12|15
茶晶方糖|12|11
白水晶长方|17|28
紫水晶双尖|8|15
粉水晶双尖|9|12
黄水晶双尖|9|23
奶体白双尖|10|10
银曜石双尖|9|13
黄塔晶双尖|8|15
白水晶双尖|8|19
红胶花双尖|10|23
深色红胶花|10|23
茶晶双尖|10|10
白水晶蝴蝶|17|10
青金石五星|14|17
柠檬晶小花|14|21
海蓝宝月亮|15|27
银曜石貔貅|20|17
海蓝宝四叶草|15|31
草莓晶爱心|12|13
海蓝宝爱心|13|42
白水晶胖星星|15|31
银曜石月球|12|15
白水晶爱心|18|31
银曜石猫猫|15|15
银曜石醒狮|14|15
茶晶蝴蝶结|15|10
粉水晶爱心|12|13
黑金超貔貅|17|31
萤石回纹珠|12|10
银曜石猫爪|15|15
银耀石貔貅|17|13
白水晶小熊|17|15
银曜石胖星星|13|15
白水晶花托|16|15
黑曜石爱心|14|15
粉晶小花|16|13
黄水晶切面珠 大|15|37
黄水晶弧形切面珠|18|47
白水晶钻切|10|13
白水晶椭圆切面珠|14|17
白水晶圆柱切面|16|31
茶晶切面珠|15|21
白水晶弧形切面珠|12|17
黄水晶切面珠|12|21
""",
    "随型",
    "accessory",
    False,
)
SHAPED_SWITCHABLE = rows(
    """
白水晶桶珠|4|6
薰衣草方糖|10|16
柠檬晶方糖|8|9
白水晶方糖|11|13
奶体白方糖|10|11
白水晶方糖对角|12|38
白水晶切面珠|8|5
""",
    "随型",
    "accessory",
    True,
)

# Accessories were read from the same live UI.  Only entries whose card image is
# among the 235 local files will be written; the source list contains additional
# cards that deliberately have no local image counterpart yet.
ACCESSORIES_FIXED = rows(
    """
千伞流苏吊坠|8|15|
锆石蝴蝶吊坠银|10|6|适合7m至以上
半蝴蝶银|2|7|适合10m至以上
镂空叶子吊坠|8|3|适合7m-13m
岁岁平安|4|5|
小金条吊坠|28|6|
金色随型吊坠|9|5|
银色锥形吊坠|16|4|
金色锥形吊坠|16|4|
小金砖吊坠|15|15|
银色随型吊坠|9|5|
花型吊坠|15|6|
多叶吊坠|4|3.5|适合8m至以上
古荷吊坠|16|7|适合8m以上
半蝴蝶金|2|7|适合10m至以上
锆石蝴蝶吊坠金|10|6|适合7m至以上
蝴蝶珍珠|8|5|适合7m至以上
棱形锆石隔环银|7|3.5|适合9m-14m
四方隔珠金|10|3|适合10m至以上
藏银多面隔珠|10|2.5|
银色随型隔片|10|2|
藏银双花|9|2.5|
金色猫猫|19|6|
金色随型隔片|10|2|
金色花饰|13|6|
镂空金丝球|8|1.5|适合6m-10m
淡紫锆石小隔片|2|2|
镂空银丝球|8|1.5|适合6m-10m
小隔珠银|6|1|
隔珠银|9|6|适合6m-11m
锆石长隔环银|9|6|适合9m-14m
四方隔珠银|10|2.5|适合10m至以上
爱心隔环金|7|3|适合8m-11m
爱心隔环银|7|3|适合8m-11m
长锆石隔环银|7|3.5|适合10m至以上
花朵长隔环银|9|6|适合7m-12m
隔珠金|9|6.5|适合6m-11m
四环绕戒隔环银|8|3.5|适合12m至以上
四环绕隔环银|8|8|适合10m至以上
单排珍珠隔环金|6|3.5|适合10m至以上
花嵌隔环银|10|3|适合10m-13m
花嵌隔环金|10|3|适合10m-13m
五叶饶戒|13|5|适合10m-13m
单排珍珠隔环银|6|3.5|适合10m至以上
花朵长隔环金|9|6|适合7m-12m
荆棘隔环金|8|5|适合11m至以上
荆棘隔环银|8|5|适合11m至以上
金棱隔环|7|5|适合11m至以上
镂空隔环戒银|5|3|适合8m至以上
细镂空隔环银|5|3|适合10m至以上
细镂空隔环金|5|3|适合10m至以上
双排珍珠隔环|5|2|适合8m-11m
锆石长隔环金|9|5|适合9m-14m
四环绕隔环金|8|8|适合10m至以上
四环绕戒隔环|8|3.5|适合12m至以上
双戒环绕金|12|4|适合11m至以上
双戒环绕银|12|4|适合11m至以上
镂空隔环银|8|3|适合9m-12m
水棱隔珠|9|3|适合11m-13m
小隔珠金|6|1|
古银花托|8|1.5|适合9m至以上
镂空隔珠金|9|3|
藏银隔珠|9|2|
镂空隔珠银|9|3|
锆石蝴蝶银|11|4.5|
蝴蝶配饰金|15|4.5|适合8m-15m
月蝶锆石隔珠|9|3|适合7m-12m
锆石隔环金|5|3.5|适合10m-14m
锆石隔环|5|4.5|适合9m-12m
古银隔片|3|2|适合9m至以上
水滴锆石金|12|3.5|适合6m-12
双花隔环金|6|3|适合11m至以上
锆石蝴蝶金|11|4.5|
通用小隔珠银|3|1|
古银方隔环|4|1.5|
藏银隔环戒|4|4|适合9m至以上
棱形锆石隔环金|7|4.5|适合9m-14m
藏银多环隔珠|7|2.5|
藏银弧形隔珠|10|3|
藏银花隔珠|6|2|
藏银隔环|6|1.5|
藏银花隔环|5|2.5|
通用小隔珠金|3|1|
藏银棱形隔珠|9|2.5|
金色厚隔片|2|1.5|
贝珠|10|6|
琥珀隔片|4|6|
藏银环隔片|4|1.5|
藏银环隔珠|6|2|
檀木隔片厚|4|2|
花型魔盒金|9|4|适合6m-12m
雪花魔盒|14|5|适合6m至以上
椭圆锆石魔盒金|12|4|适合6m至以上
葫芦魔盒奶黄|13|4|适合6m-11m
花型魔盒银|9|4|适合6m-12m
葫芦魔盒黄|13|4|适合6m-11m
葫芦魔盒蓝|13|4|适合6m-11m
千层花银|14|8|适合6m至以上
千层花金|14|8|适合6m至以上
四叶草（可转动）|14|8|适合6m至以上
粉水晶五角星|19|22|
黄水晶五角星|19|22|
花瓣锆石花托左|4|4|适合12m-15m
锆石花托左|4|5|适合13m至以上
多角花托金右|5|5|适合12m以上
多角花托银右|5|5|适合12m以上
梅花花托左|3|5|适合9m-12m
透明白跑环|6|13|
檀木跑环|6|13|
冰曜石跑环|5|13|
海蓝宝跑环|5|22|
灰月光跑环|5|24|
红玛瑙跑环|5|7|
草莓晶跑环|5|13|
蓝月光跑环|5|18|
银曜石跑环|6|15|
""",
    "配饰",
    "accessory",
    False,
)
ACCESSORIES_SWITCHABLE = rows(
    """
金色小方块|6|2.5|
银色小方块|6|2.5|
沉香方珠|7|3|
银色弧形隔片|2|1|适合6m-10m
锆石隔环银|3|2|适合10m-14m
花嵌隔环金|10|3|适合10m-13m
金色弧形隔片|2|1|适合6m-10m
镂空隔环金|5|2.5|适合8m至以上
银色隔片|2|1|
藏银叶子隔珠|10|2.5|
金色隔片|2|1|
藏银祥云隔珠|11|3|
藏银绣球花|11|3|
大锆石魔盒|14|6.5|适合7m及以上
白锆魔盒|10|4.7|适合6m-12m
紫锆魔盒|10|6|适合6m-12m
小金花花托右|3|1|适合9m至以上
小银花花托右|3|1|适合9m至以上
紫水晶跑环|5|20|
黑曜石跑环|5|13|
贝珠跑环|5|13|
""",
    "配饰",
    "accessory",
    True,
)

# Full -/+ boundary scans performed in the running mini-program.  Keeping these
# separate from the default-card list makes every promoted "complete" Variant
# traceable to a real interaction rather than to an assumed standard size set.
COMPLETE_BOUNDARY_VARIANTS = [
    # 2026-07-21 UI: searched within 配饰; both − and + left the 14 mm card unchanged.
    {"name": "大锆石魔盒", "size": 14.0, "price": 6.5, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": "适合7m及以上"},
    {"name": "金色小方块", "size": 6.0, "price": 2.5, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": ""},
    {"name": "金色小方块", "size": 8.0, "price": 4.0, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": ""},
    {"name": "银色小方块", "size": 6.0, "price": 2.5, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": ""},
    {"name": "银色小方块", "size": 8.0, "price": 4.0, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": ""},
    {"name": "沉香方珠", "size": 7.0, "price": 3.0, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": ""},
    {"name": "沉香方珠", "size": 8.0, "price": 4.0, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": ""},
    {"name": "沉香方珠", "size": 9.0, "price": 5.0, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": ""},
    {"name": "银色弧形隔片", "size": 2.0, "price": 1.0, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": "适合6m-10m"},
    {"name": "银色弧形隔片", "size": 3.0, "price": 1.0, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": "适合6m-10m"},
    {"name": "金色弧形隔片", "size": 2.0, "price": 1.0, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": "适合6m-10m"},
    {"name": "金色弧形隔片", "size": 3.0, "price": 1.0, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": "适合6m-10m"},
    {"name": "锆石隔环银", "size": 3.0, "price": 2.0, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": "适合10m-14m"},
    {"name": "锆石隔环银", "size": 5.0, "price": 3.0, "source_tab": "配饰", "component": "accessory", "switchable": True, "complete": True, "fit": "适合10m-14m"},
]


def header_index(ws) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in ws[1]}


def as_number(value: float) -> int | float:
    return int(value) if value.is_integer() else value


def main() -> None:
    if not BOOK.exists():
        raise FileNotFoundError(BOOK)
    if not IMAGES.exists():
        raise FileNotFoundError(IMAGES)

    observed = (
        CRYSTALS + NATURAL_SWITCHABLE + NATURAL_FIXED + SHAPED_FIXED + SHAPED_SWITCHABLE
        + ACCESSORIES_FIXED + ACCESSORIES_SWITCHABLE + COMPLETE_BOUNDARY_VARIANTS
    )
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for item in observed:
        grouped[norm(str(item["name"]))].append(item)

    workbook = load_workbook(BOOK)
    summary_ws, master_ws, variant_ws = workbook.worksheets[:3]
    master_h = header_index(master_ws)
    variant_h = header_index(variant_ws)

    image_names: dict[str, list[str]] = defaultdict(list)
    for path in IMAGES.iterdir():
        if path.is_file() and path.suffix.lower() in {".png", ".webp", ".jpg", ".jpeg"}:
            image_names[norm(path.stem)].append(path.stem)

    master_rows: dict[str, list[int]] = defaultdict(list)
    for row in range(2, master_ws.max_row + 1):
        master_rows[norm(str(master_ws.cell(row, master_h["商品名称_文件名"]).value or ""))].append(row)

    existing_source_rows = {
        str(variant_ws.cell(row, variant_h["来源VariantID"]).value or ""): row
        for row in range(2, variant_ws.max_row + 1)
    }
    written_products = 0
    written_variants = 0
    unmatched_observed: list[str] = []

    for key, items in grouped.items():
        source_name = str(items[0]["name"])
        image_key = norm(APP_NAME_TO_IMAGE_NAME.get(source_name, source_name))
        is_explicit_alias = image_key != key
        if image_key not in image_names or len(image_names[image_key]) != 1 or len(master_rows[image_key]) != 1:
            unmatched_observed.append(str(items[0]["name"]))
            continue
        row = master_rows[image_key][0]
        material_key = str(master_ws.cell(row, master_h["研究商品键"]).value)
        name = str(items[0]["name"])
        source_tab = str(items[0]["source_tab"])
        component = str(items[0]["component"])
        fit = str(items[0].get("fit") or "")
        by_size: dict[float, float] = {}
        for item in items:
            by_size[float(item["size"])] = float(item["price"])
        complete_sizes = {
            float(item["size"])
            for item in items
            if bool(item["complete"])
        }
        # Default-card observations may coexist with later full boundary checks.
        # The complete observations supersede those duplicate default readings.
        all_complete = set(by_size) == complete_sizes
        size_values = sorted(by_size)
        price_values = [by_size[size] for size in size_values]
        pending = "" if all_complete else "needs-boundary-scan"

        values = {
            "组件类型": component,
            "分类": source_tab,
            "匹配状态": "mini-program-ui-explicit-alias" if is_explicit_alias else "mini-program-ui-exact",
            "证据等级": EVIDENCE,
            "对应数据商品名": name,
            "Variant数": len(size_values),
            "尺寸列表_mm": ", ".join(str(as_number(value)) for value in size_values),
            "价格列表_CNY": ", ".join(str(as_number(value)) for value in price_values),
            "缺尺寸行": pending,
            "缺价格行": pending,
            "导入状态": "draft-observed",
            "本地交互状态": "mini-program-ui-observed",
            "生产状态": "not-for-production",
        }
        for header, value in values.items():
            master_ws.cell(row, master_h[header], value)
        written_products += 1

        for size, price in by_size.items():
            source_id = f"LINGANSHI-APP-{OBSERVED_ON}:{source_tab}:{name}:{as_number(size)}"
            complete_note = "all UI size bounds checked" if all_complete else "UI default observed; size bounds still need - then + boundary scan"
            if source_id in existing_source_rows:
                variant_ws.cell(existing_source_rows[source_id], variant_h["说明"], complete_note)
                variant_ws.cell(existing_source_rows[source_id], variant_h["适配珠径"], fit)
                continue
            new_row = {
                "研究商品键": material_key,
                "商品名称_文件名": image_names[image_key][0],
                "对应数据商品名": name,
                "匹配状态": "mini-program-ui-explicit-alias" if is_explicit_alias else "mini-program-ui-exact",
                "证据等级": EVIDENCE,
                "Variant键": f"{material_key}-app-{as_number(size)}",
                "来源VariantID": source_id,
                "尺寸_mm": as_number(size),
                "价格_CNY": price,
                "来源分类": source_tab,
                "来源库存": "",
                "来源图片URL": "",
                "重量_g": "",
                "2D占位_mm": "",
                "方向规则": "",
                "适配珠径": fit,
                "导入状态": "draft-observed",
                "说明": complete_note,
            }
            variant_ws.append([new_row[header] for header in variant_h])
            existing_source_rows[source_id] = variant_ws.max_row
            written_variants += 1

    audited_without_source = 0
    for image_name, note in UNRESOLVED_IMAGE_AUDIT.items():
        key = norm(image_name)
        if len(master_rows[key]) != 1:
            raise ValueError(f"Expected exactly one master row for unresolved image: {image_name}")
        row = master_rows[key][0]
        values = {
            "组件类型": "accessory",
            "分类": "source-audit-pending",
            "匹配状态": "mini-program-ui-no-safe-match",
            "证据等级": EVIDENCE,
            "对应数据商品名": "",
            "Variant数": "",
            "尺寸列表_mm": "",
            "价格列表_CNY": "",
            "缺尺寸行": "requires-current-source-match",
            "缺价格行": "requires-current-source-match",
            "导入状态": "blocked-source-match",
            "本地交互状态": "mini-program-ui-audited",
            "生产状态": "not-for-production",
        }
        for header, value in values.items():
            master_ws.cell(row, master_h[header], value)
        # Preserve the evidence in an existing main-workbook column rather than
        # creating another worksheet or external audit table.
        master_ws.cell(row, master_h["Live候选名"], note)
        audited_without_source += 1

    metric_rows = [
        (f"{OBSERVED_ON} 小程序实测精确产品", written_products, "written to master workbook"),
        (f"{OBSERVED_ON} 小程序实测Variant行", written_variants, "written to Variant audit"),
        (f"{OBSERVED_ON} 已读卡但无图片精确匹配", len(unmatched_observed), "kept as no-write; requires name review"),
        (f"{OBSERVED_ON} 图片商品已核验但无安全来源", audited_without_source, "blocked; no price/size substitution"),
    ]
    existing_metrics = {
        str(summary_ws.cell(row, 1).value or ""): row
        for row in range(2, summary_ws.max_row + 1)
    }
    for metric, amount, status in metric_rows:
        if metric in existing_metrics:
            summary_ws.cell(existing_metrics[metric], 2, amount)
            summary_ws.cell(existing_metrics[metric], 3, status)
            continue
        summary_ws.append([metric, amount, status])

    workbook.save(BOOK)
    print(f"master_products_written={written_products}")
    print(f"variant_rows_written={written_variants}")
    print(f"observed_without_exact_image_match={len(unmatched_observed)}")
    print("unmatched=" + " | ".join(unmatched_observed[:40]))


if __name__ == "__main__":
    main()
