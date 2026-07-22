from __future__ import annotations

import csv
import difflib
import hashlib
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
RESEARCH = ROOT / "data" / "v3" / "research"
IMAGE_SOURCE = Path(r"C:\Users\Dylan\Desktop\灵感石验室")
LEGACY_WORKBOOK = RESEARCH / "linganshi-material-catalog-20260715.xlsx"
LIVE_JSON = RESEARCH / "linganshi-live-materials-20260720.json"
REVIEW_WORKBOOK = RESEARCH / "linganshi-image-mapping-review-20260720.xlsx"
DRAFT_IMPORT = RESEARCH / "linganshi-235-draft-catalog-20260720.import.csv"
MEDIA_MANIFEST = RESEARCH / "linganshi-235-media-upload-manifest-20260720.csv"

IMPORT_HEADERS = [
    "material_key", "component_type", "library_tab_slug", "category_slug", "name_en", "primary_color",
    "color_tags", "intention_tags", "material_image_url", "material_status",
    "material_sort_order", "variant_key", "size_mm", "shape", "price", "weight_g",
    "occupied_length_mm", "display_scale", "variant_image_url", "stock_status",
    "stock_quantity", "compatibility", "compatible_bead_sizes", "orientation_mode",
    "mirrored_variant_key", "allowed_orientations", "allowed_positions",
    "neighbor_constraints", "variant_status", "variant_sort_order", "source_name",
]

# These aliases only bridge an old source-image name to a current card whose material and
# physical shape remain explicit in both names. They do not merge the source materials.
EXPLICIT_ALIASES = {
    "冰川蓝海蓝宝": "冰种海蓝宝",
    "柠檬黄水晶": "透体柠檬黄水晶",
    "蓝天白云海蓝宝": "蓝天白云",
    "薰衣草": "薰衣草紫水晶",
    "白水晶双尖": "白水晶双尖柱",
    "白水晶方糖": "白水晶刻面方糖",
    "白水晶爱心": "白水晶爱心雕刻件",
    "粉水晶双尖": "粉水晶双尖柱",
    "粉水晶五角星": "粉五角星",
    "透明白跑环": "白水晶跑环",
}

DECOR_HINTS = (
    "吊坠", "隔环", "隔珠", "隔片", "花托", "魔盒", "绕戒", "跑环", "配饰",
    "金丝球", "银丝球", "小金条", "小金砖", "猫猫", "锥形", "流苏", "珍珠",
)


def stable_key(name: str) -> str:
    return "linganshi-" + hashlib.sha1(name.encode("utf-8")).hexdigest()[:12]


def numeric_size(value: Any) -> float:
    text = str(value or "").strip().replace("..", ".")
    match = re.search(r"\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else 0.0


def numeric_price(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def format_number(value: float) -> str:
    return f"{value:.4f}".rstrip("0").rstrip(".") or "0"


def component_type(name: str, category: str) -> str:
    if category in {"吊坠", "隔珠", "水晶配饰"} or any(hint in name for hint in DECOR_HINTS):
        return "accessory"
    return "crystal"


def category_slug(component: str, category: str) -> str:
    if component == "accessory":
        if category == "吊坠" or "吊坠" in category:
            return "charm"
        return "accessory"
    if category in {"文玩", "和田玉", "玛瑙", "曜石", "其他"}:
        return "natural-stone"
    return "crystal"


def style_sheet(sheet, table_name: str) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if sheet.max_row > 1:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in sheet.columns:
        letter = column[0].column_letter
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[letter].width = min(max(width, 10), 52)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]], table_name: str) -> None:
    use_blank_active = len(workbook.sheetnames) == 1 and workbook.active.max_row == 1 and workbook.active["A1"].value is None
    sheet = workbook.active if use_blank_active else workbook.create_sheet()
    sheet.title = title
    headers = list(rows[0]) if rows else ["说明"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column, value=str(header))
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    style_sheet(sheet, table_name)


def load_legacy_variants() -> tuple[dict[str, list[dict[str, Any]]], dict[str, dict[str, str]]]:
    workbook = load_workbook(LEGACY_WORKBOOK, read_only=True, data_only=True)
    sheet = workbook["变体明细"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(rows)]
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    metadata: dict[str, dict[str, str]] = {}
    for values in rows:
        row = dict(zip(headers, values))
        name = str(row.get("商品名称") or "").strip()
        if not name:
            continue
        by_name[name].append(row)
        metadata.setdefault(name, {
            "一级类目": str(row.get("一级类目") or "").strip(),
            "二级类目": str(row.get("二级类目") or "").strip(),
        })
    return by_name, metadata


def load_live_variants() -> dict[str, list[dict[str, Any]]]:
    payload = json.loads(LIVE_JSON.read_text(encoding="utf-8"))
    materials = payload.get("data", {}).get("materials", [])
    if not isinstance(materials, list):
        raise ValueError("Live material payload has no data.materials array")
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in materials:
        if not isinstance(row, dict):
            continue
        name = str(row.get("name") or "").strip()
        if name:
            by_name[name].append(row)
    return by_name


def candidate_names(name: str, live_names: list[str]) -> str:
    matches = difflib.get_close_matches(name, live_names, n=3, cutoff=0.15)
    return " | ".join(matches)


def image_metadata(path: Path) -> dict[str, Any]:
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    with Image.open(path) as image:
        width, height = image.size
        detected = str(image.format or "unknown").lower()
        alpha = image.mode in {"RGBA", "LA"} or "transparency" in image.info
    return {
        "图片文件名": path.name,
        "图片绝对路径": str(path),
        "SHA256": digest,
        "宽_px": width,
        "高_px": height,
        "透明通道": "yes" if alpha else "no",
        "真实格式": detected,
    }


def legacy_variant_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        result.append({
            "source_id": "",
            "source_name": "legacy-verified-workbook-20260715",
            "category": str(row.get("二级类目") or row.get("一级类目") or "").strip(),
            "size_mm": numeric_size(row.get("尺寸_mm")),
            "price": numeric_price(row.get("单价_CNY")),
            "image_url": "",
            "stock": None,
            "image_metrics": {},
        })
    return result


def live_variant_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        result.append({
            "source_id": str(row.get("id") or ""),
            "source_name": "live-mini-program-api-20260720",
            "category": str(row.get("category") or "").strip(),
            "size_mm": numeric_size(row.get("diameter") or row.get("horizontalLength")),
            "price": numeric_price(row.get("price")),
            "image_url": str(row.get("imageUrl") or row.get("gridImageUrl") or "").strip(),
            "stock": row.get("stock"),
            "image_metrics": row.get("imageMetrics") if isinstance(row.get("imageMetrics"), dict) else {},
        })
    return result


def dedupe_variants(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Keep one deterministic card Variant per displayed size.

    The current mini-program endpoint contains a few duplicate records for the same
    product and size. T17 requires unique Variant keys and the size selector must not
    display duplicate steps, so prefer a positive-price/in-stock row and keep source
    order as the final tie-breaker.
    """
    by_size: dict[float, dict[str, Any]] = {}
    for row in rows:
        size = float(row.get("size_mm") or 0)
        current = by_size.get(size)
        if current is None:
            by_size[size] = row
            continue
        row_stock = row.get("stock")
        current_stock = current.get("stock")
        row_score = (
            numeric_price(row.get("price")) > 0,
            not isinstance(row_stock, (int, float)) or row_stock > 0,
        )
        current_score = (
            numeric_price(current.get("price")) > 0,
            not isinstance(current_stock, (int, float)) or current_stock > 0,
        )
        if row_score > current_score:
            by_size[size] = row
    return [by_size[size] for size in sorted(by_size)]


def main() -> None:
    for required in (IMAGE_SOURCE, LEGACY_WORKBOOK, LIVE_JSON):
        if not required.exists():
            raise FileNotFoundError(required)
    for output in (REVIEW_WORKBOOK, DRAFT_IMPORT, MEDIA_MANIFEST):
        if ROOT not in output.parents:
            raise RuntimeError(f"Output escaped project root: {output}")

    images = sorted(IMAGE_SOURCE.glob("*.webp"), key=lambda path: path.name)
    if len(images) != 235:
        raise RuntimeError(f"Expected 235 source images, found {len(images)}")
    if len({path.stem for path in images}) != 235:
        raise RuntimeError("Image base names must be unique product names")

    legacy_by_name, legacy_metadata = load_legacy_variants()
    live_by_name = load_live_variants()
    live_names = sorted(live_by_name)
    existing_media: dict[str, dict[str, str]] = {}
    if MEDIA_MANIFEST.is_file():
        with MEDIA_MANIFEST.open(encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                key = str(row.get("material_key") or "").strip()
                if key:
                    existing_media[key] = row

    product_rows: list[dict[str, Any]] = []
    review_variant_rows: list[dict[str, Any]] = []
    import_rows: list[dict[str, Any]] = []
    media_rows: list[dict[str, Any]] = []
    mapping_counts: dict[str, int] = defaultdict(int)
    shared_hashes: dict[str, list[str]] = defaultdict(list)
    image_records: list[tuple[Path, dict[str, Any]]] = []
    for image in images:
        meta = image_metadata(image)
        shared_hashes[str(meta["SHA256"])].append(image.name)
        image_records.append((image, meta))

    for order, (image, meta) in enumerate(image_records, start=10):
        name = image.stem
        material_key = stable_key(name)
        alias = EXPLICIT_ALIASES.get(name, "")

        if name in live_by_name:
            mapping_status = "live-exact"
            source_product_name = name
            evidence_level = "A-current-exact"
            variants = live_variant_rows(live_by_name[name])
        elif name in legacy_by_name:
            mapping_status = "legacy-exact"
            source_product_name = name
            evidence_level = "A-legacy-exact"
            variants = legacy_variant_rows(legacy_by_name[name])
        elif alias and alias in live_by_name:
            mapping_status = "live-explicit-alias"
            source_product_name = alias
            evidence_level = "B-explicit-alias-draft"
            variants = live_variant_rows(live_by_name[alias])
        else:
            mapping_status = "pending-product-data"
            source_product_name = ""
            evidence_level = "D-missing-variant-evidence"
            variants = [{
                "source_id": "",
                "source_name": "pending-review-placeholder",
                "category": legacy_metadata.get(name, {}).get("二级类目", ""),
                "size_mm": 0.0,
                "price": 0.0,
                "image_url": "",
                "stock": None,
                "image_metrics": {},
            }]

        variants = dedupe_variants(variants)

        mapping_counts[mapping_status] += 1
        category = next((str(row.get("category") or "").strip() for row in variants if row.get("category")), "")
        if not category:
            category = legacy_metadata.get(name, {}).get("二级类目", "") or legacy_metadata.get(name, {}).get("一级类目", "")
        ctype = component_type(name, category)
        cslug = category_slug(ctype, category)
        positive_sizes = [row["size_mm"] for row in variants if row["size_mm"] > 0]
        positive_prices = [row["price"] for row in variants if row["price"] > 0]
        missing_price_rows = sum(1 for row in variants if row["price"] <= 0)
        missing_size_rows = sum(1 for row in variants if row["size_mm"] <= 0)
        product_rows.append({
            "研究商品键": material_key,
            "商品名称_文件名": name,
            **meta,
            "组件类型": ctype,
            "分类": category,
            "匹配状态": mapping_status,
            "证据等级": evidence_level,
            "对应数据商品名": source_product_name,
            "Variant数": len(variants),
            "尺寸列表_mm": " | ".join(format_number(value) for value in positive_sizes),
            "价格列表_CNY": " | ".join(format_number(value) for value in positive_prices),
            "缺尺寸行": missing_size_rows,
            "缺价格行": missing_price_rows,
            "Live候选名": candidate_names(name, live_names),
            "共用图片": " | ".join(shared_hashes[str(meta["SHA256"])]) if len(shared_hashes[str(meta["SHA256"])]) > 1 else "",
            "导入状态": "draft-only",
            "本地交互状态": "enabled-draft-review" if missing_size_rows == 0 and missing_price_rows == 0 else "disabled-pending-specs",
            "生产状态": "blocked-until-english-name-weight-occupancy-and-rules-reviewed",
        })

        uploaded_media = existing_media.get(material_key, {})
        material_image_url = str(uploaded_media.get("wordpress_media_url") or "").strip()
        media_rows.append({
            "material_key": material_key,
            "source_product_name_zh": name,
            "source_image_path": str(image),
            "upload_filename": material_key + ".webp",
            "wordpress_attachment_id": str(uploaded_media.get("wordpress_attachment_id") or "").strip(),
            "wordpress_media_url": material_image_url,
            "upload_status": str(uploaded_media.get("upload_status") or "pending-wordpress-media-upload").strip(),
        })

        for variant_order, variant in enumerate(variants, start=10):
            size = float(variant["size_mm"])
            price = float(variant["price"])
            suffix = format_number(size).replace(".", "p") if size > 0 else f"pending-{variant_order}"
            variant_key = f"{material_key}-{suffix}mm"
            stock = variant.get("stock")
            stock_quantity = int(stock) if isinstance(stock, (int, float)) and stock >= 0 else ""
            stock_status = "instock" if stock is None or not isinstance(stock, (int, float)) or stock != 0 else "outofstock"
            occupied = size if size > 0 else 0.0
            weight = 0.0
            review_variant_rows.append({
                "研究商品键": material_key,
                "商品名称_文件名": name,
                "对应数据商品名": source_product_name,
                "匹配状态": mapping_status,
                "证据等级": evidence_level,
                "Variant键": variant_key,
                "来源VariantID": variant.get("source_id", ""),
                "尺寸_mm": size,
                "价格_CNY": price,
                "来源分类": variant.get("category", ""),
                "来源库存": "" if stock is None else stock,
                "来源图片URL": variant.get("image_url", ""),
                "重量_g": weight,
                "2D占位_mm": occupied,
                "方向规则": "pending" if ctype == "accessory" else "none",
                "适配珠径": "pending" if ctype == "accessory" else "",
                "导入状态": "draft",
                "说明": "重量、最终占位、英文名和配件规则未审核前不得改为 live",
            })
            import_rows.append({
                "material_key": material_key,
                "component_type": ctype,
                "library_tab_slug": ctype,
                "category_slug": cslug,
                # Source Chinese names are intentionally preserved only for draft admin review.
                "name_en": name,
                "primary_color": "",
                "color_tags": "",
                "intention_tags": "",
                "material_image_url": material_image_url,
                "material_status": "draft",
                "material_sort_order": order,
                "variant_key": variant_key,
                "size_mm": format_number(size),
                "shape": "round" if ctype == "crystal" else "accessory",
                "price": format_number(price),
                "weight_g": "0",
                "occupied_length_mm": format_number(occupied),
                "display_scale": "1",
                "variant_image_url": material_image_url,
                "stock_status": stock_status,
                "stock_quantity": stock_quantity,
                "compatibility": "crystal,elastic-cord" if ctype == "crystal" else "accessory,elastic-cord",
                "compatible_bead_sizes": "",
                "orientation_mode": "none",
                "mirrored_variant_key": "",
                "allowed_orientations": "",
                "allowed_positions": "",
                "neighbor_constraints": "",
                "variant_status": "draft",
                "variant_sort_order": variant_order,
                "source_name": f"Linganshi research draft; {variant['source_name']}; source product={source_product_name or name}",
            })

    summary_rows = [
        {"指标": "图片商品主记录", "数量": len(product_rows), "状态": "235/235 已建 draft 映射记录"},
        {"指标": "当前接口完全同名", "数量": mapping_counts["live-exact"], "状态": "使用 2026-07-20 当前 Variant"},
        {"指标": "旧表完全同名", "数量": mapping_counts["legacy-exact"], "状态": "使用此前逐规格核验 Variant"},
        {"指标": "显式别名", "数量": mapping_counts["live-explicit-alias"], "状态": "仅 draft；保留独立商品，不合并"},
        {"指标": "缺产品数据", "数量": mapping_counts["pending-product-data"], "状态": "生成 0 值 draft 占位，不得上架"},
        {"指标": "导入 Variant 行", "数量": len(import_rows), "状态": "全部 draft；含真实 Variant 与待审占位"},
        {"指标": "缺重量的 Variant", "数量": sum(1 for row in review_variant_rows if row["重量_g"] <= 0), "状态": "阻止 live"},
        {"指标": "缺尺寸的 Variant", "数量": sum(1 for row in review_variant_rows if row["尺寸_mm"] <= 0), "状态": "阻止 live"},
        {"指标": "缺价格的 Variant", "数量": sum(1 for row in review_variant_rows if row["价格_CNY"] <= 0), "状态": "阻止 live"},
        {"指标": "图片共用例外", "数量": sum(1 for row in product_rows if row["共用图片"]), "状态": "白水晶与高品净体白水晶仍为独立商品"},
    ]
    selectable_product_count = sum(
        1 for row in product_rows if row["缺尺寸行"] == 0 and row["缺价格行"] == 0
    )
    summary_rows.insert(1, {
        "指标": "本地可选商品",
        "数量": selectable_product_count,
        "状态": "有尺寸和价格证据；仍为 draft，不等于可上架",
    })
    summary_rows.insert(2, {
        "指标": "待审核商品",
        "数量": len(product_rows) - selectable_product_count,
        "状态": "缺商品规格或价格；仅显示映射卡，不可加入设计",
    })

    live_rows = []
    for live_name in live_names:
        for row in live_variant_rows(live_by_name[live_name]):
            live_rows.append({
                "当前商品名": live_name,
                "来源VariantID": row["source_id"],
                "分类": row["category"],
                "尺寸_mm": row["size_mm"],
                "价格_CNY": row["price"],
                "库存": "" if row["stock"] is None else row["stock"],
                "图片URL": row["image_url"],
                "证据来源": "GET /api/workbench/materials @ 2026-07-20",
            })

    workbook = Workbook()
    write_sheet(workbook, "映射摘要", summary_rows, "MappingSummary")
    write_sheet(workbook, "235商品主表", product_rows, "ProductMaster235")
    write_sheet(workbook, "导入Variant审核", review_variant_rows, "ImportVariantReview")
    write_sheet(workbook, "当前接口454行", live_rows, "CurrentLiveVariants")
    write_sheet(workbook, "媒体上传清单", media_rows, "MediaUploadManifest")
    workbook.save(REVIEW_WORKBOOK)

    with DRAFT_IMPORT.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=IMPORT_HEADERS)
        writer.writeheader()
        writer.writerows(import_rows)
    with MEDIA_MANIFEST.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(media_rows[0]))
        writer.writeheader()
        writer.writerows(media_rows)

    checked = load_workbook(REVIEW_WORKBOOK, read_only=True, data_only=True)
    print(REVIEW_WORKBOOK)
    print(" | ".join(f"{sheet.title}:{sheet.max_row - 1}" for sheet in checked.worksheets))
    print(DRAFT_IMPORT, f"rows={len(import_rows)}")
    print(MEDIA_MANIFEST, f"rows={len(media_rows)}")
    print("mapping=" + json.dumps(mapping_counts, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
