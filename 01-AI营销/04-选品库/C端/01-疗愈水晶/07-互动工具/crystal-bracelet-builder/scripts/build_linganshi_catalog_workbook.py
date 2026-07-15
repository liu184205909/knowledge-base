from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(r"D:\Code\knowledge-base\01-AI营销\04-选品库\C端\01-疗愈水晶\07-互动工具\crystal-bracelet-builder")
RESEARCH = ROOT / "data" / "v3" / "research"
MANIFEST = RESEARCH / "linganshi-category-manifest-20260715.csv"
WHITE_VARIANTS = RESEARCH / "domestic-material-variants-linganshi-20260714.csv"
AUDIT_VARIANTS = RESEARCH / "linganshi-variant-audit-20260715.csv"
OUTPUT = RESEARCH / "linganshi-material-catalog-20260715.xlsx"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream))


def add_table(ws, name: str) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row > 1 and ws.max_column > 0:
        table = Table(displayName=name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in ws.columns:
        letter = column[0].column_letter
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[letter].width = min(max(width, 10), 42)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_sheet(ws, rows: list[dict[str, object]], headers: list[str], table_name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    add_table(ws, table_name)


def main() -> None:
    manifest = read_csv(MANIFEST)
    variant_rows: list[dict[str, object]] = []

    for row in read_csv(WHITE_VARIANTS):
        variant_rows.append({
            "一级类目": "水晶", "二级类目": "白水晶",
            "商品名称": row["product_name_original"], "尺寸_mm": row["size_mm"],
            "单价_CNY": row["price_visible"], "币种": row["currency"],
            "规格核验状态": row["variant_coverage_status"], "已核验规格数": row["variant_count_verified"],
            "图片映射状态": row["image_local_mapping_status"], "证据文件": WHITE_VARIANTS.name,
            "备注": row["notes"],
        })

    for row in read_csv(AUDIT_VARIANTS):
        variant_rows.append({
            "一级类目": row["category_level_1"], "二级类目": row["category_level_2"],
            "商品名称": row["product_name_original"],
            "尺寸_mm": row["size_mm"], "单价_CNY": row["price_visible"], "币种": row["currency"],
            "规格核验状态": row["variant_coverage_status"], "已核验规格数": row["variant_count_verified"],
            "图片映射状态": row["image_local_mapping_status"], "证据文件": AUDIT_VARIANTS.name,
            "备注": row["notes"],
        })

    counts = defaultdict(lambda: {"products": set(), "variants": 0})
    for row in variant_rows:
        key = (row["一级类目"], row["二级类目"])
        counts[key]["products"].add(row["商品名称"])
        counts[key]["variants"] += 1

    reconciliation = []
    for row in sorted(manifest, key=lambda x: (x["category_level_1"], int(x["display_order"]))):
        key = (row["category_level_1"], row["category_level_2"])
        stat = counts[key]
        reconciliation.append({
            "一级类目": key[0], "二级类目": key[1], "目录序号": int(row["display_order"]),
            "目录状态": row["directory_status"], "已采商品数": len(stat["products"]),
            "已采规格记录数": stat["variants"],
            "采集状态": "已完成（待图片映射）" if stat["variants"] else "待补采",
        })

    workbook = Workbook()
    directory = workbook.active
    directory.title = "类目目录"
    write_sheet(directory, [
        {"一级类目": r["category_level_1"], "二级类目": r["category_level_2"],
         "目录序号": r["display_order"], "目录状态": r["directory_status"],
         "目录证据": r["evidence_method"], "备注": r["notes"]}
        for r in manifest
    ], ["一级类目", "二级类目", "目录序号", "目录状态", "目录证据", "备注"], "CategoryManifest")

    write_sheet(workbook.create_sheet("变体明细"), variant_rows,
                ["一级类目", "二级类目", "商品名称", "尺寸_mm", "单价_CNY", "币种",
                 "规格核验状态", "已核验规格数", "图片映射状态", "证据文件", "备注"], "VariantDetails")
    write_sheet(workbook.create_sheet("采集对账"), reconciliation,
                ["一级类目", "二级类目", "目录序号", "目录状态", "已采商品数", "已采规格记录数", "采集状态"],
                "CollectionReconciliation")
    write_sheet(workbook.create_sheet("图片映射"), [{
        "图片来源目录": r"C:\Users\Dylan\Desktop\灵感石验室", "图片数量": 221,
        "已命名映射数": 0, "状态": "待执行", "说明": "仅在名称、规格、价格与目录均可核验后移动到项目研究目录",
    }], ["图片来源目录", "图片数量", "已命名映射数", "状态", "说明"], "ImageMapping")

    workbook.save(OUTPUT)
    print(OUTPUT)
    verified = load_workbook(OUTPUT, read_only=True, data_only=True)
    print(" | ".join(f"{sheet.title}:{sheet.max_row - 1}" for sheet in verified.worksheets))


if __name__ == "__main__":
    main()
