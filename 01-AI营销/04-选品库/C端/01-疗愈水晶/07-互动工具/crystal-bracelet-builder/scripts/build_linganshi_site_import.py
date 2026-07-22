"""Prepare the verified Linganshi workbook and local media for T17 REST import.

This is deliberately an operational two-file pipeline:
  1. media-manifest.csv is updated by the media uploader with WordPress URLs;
  2. catalog-import.json is produced only after every live card has its media URL.

No price, size, category, direction or compatibility value is invented here. Rows
without a matching local WebP are reported and withheld from the catalog payload.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "data" / "v3" / "research" / "linganshi-image-mapping-review-20260720.xlsx"
DEFAULT_OUTPUT = ROOT / "data" / "v3" / "site-imports"
DEFAULT_MEDIA_DIR = Path(os.environ.get("EW_T17_SOURCE_DIR", str(Path.home() / "Desktop" / "灵感石验室")))

# Workbook product name -> confirmed local WebP filename stem. These are only
# spelling/format normalizations; all other names remain unimported for review.
WORKBOOK_TO_SOURCE = {
    "冰耀石跑环": "冰曜石跑环",
    "高透冰耀石": "高透冰曜石",
    "白水晶圆柱切面珠": "白水晶圆柱切面",
    "白水晶椭圆切面珠": "白水晶椭圆切面",
    "白水晶长方糖": "白水晶长方",
    "黄塔晶随型大": "黄塔晶随型 大",
}


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def stable_key(name: str) -> str:
    return "linganshi-" + hashlib.sha1(name.encode("utf-8")).hexdigest()[:12]


def category_key(category: str) -> str:
    return "category-" + hashlib.sha1(category.encode("utf-8")).hexdigest()[:10]


def number(value: object, field: str, row_number: int) -> float:
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"row {row_number}: {field} is not numeric") from exc
    if result <= 0:
        raise ValueError(f"row {row_number}: {field} must be greater than zero")
    return result


def computed_usd_price(value: object, cny_price: object, row_number: int) -> float:
    """Read a reviewed USD value, including the workbook's fixed USD formula.

    Excel normally stores a cached formula result, but that cache is removed by
    some non-Excel editors.  The reviewed sheet uses one explicit conversion
    formula, so evaluate only that exact rule when the cache is unavailable.
    """
    raw = text(value)
    if raw.startswith("="):
        normalized = re.sub(r"\s+", "", raw).upper()
        if re.fullmatch(r"=ROUNDUP\(D\d+\*4\.5/6\.5,0\)", normalized):
            return float(math.ceil(number(cny_price, "CNY price", row_number) * 4.5 / 6.5))
        raise ValueError(f"row {row_number}: unsupported USD formula {raw!r}")
    return number(value, "USD price", row_number)


def quantity(value: float) -> str:
    return str(int(value)) if value.is_integer() else ("%.3f" % value).rstrip("0").rstrip(".")


def normalized_direction(value: str) -> str:
    value = text(value).replace("-", "_")
    if value not in {"none", "tangent", "radial_out", "radial_in"}:
        raise ValueError(f"Unsupported direction_rule: {value!r}")
    return value


def parse_compatibility(value: str, crystal_sizes: list[float]) -> str:
    value = text(value)
    if not value:
        return ""
    numbers = [float(item) for item in re.findall(r"\d+(?:\.\d+)?", value)]
    if not numbers:
        raise ValueError(f"Cannot parse compatibility: {value!r}")
    if "以上" in value or ">=" in value:
        allowed = [size for size in crystal_sizes if size >= numbers[0]]
    elif "以下" in value or "<=" in value:
        allowed = [size for size in crystal_sizes if size <= numbers[0]]
    elif len(numbers) >= 2 and ("-" in value or "~" in value or "至" in value):
        low, high = min(numbers[0], numbers[1]), max(numbers[0], numbers[1])
        allowed = [size for size in crystal_sizes if low <= size <= high]
    else:
        allowed = [size for size in crystal_sizes if any(abs(size - wanted) < 0.001 for wanted in numbers)]
    return ",".join(quantity(size) for size in allowed)


def image_stems(media_dir: Path) -> dict[str, Path]:
    images: dict[str, Path] = {}
    for path in sorted(media_dir.glob("*.webp")):
        images[path.stem] = path
    return images


def load_rows(workbook: Path) -> list[dict[str, object]]:
    # Keep formulas visible so the fixed audited USD formula can be evaluated
    # even when an editor has not persisted Excel's cached result.
    wb = load_workbook(workbook, read_only=True, data_only=False)
    # The review workbook intentionally keeps this operational sheet third;
    # using index avoids source encoding ambiguity in non-UTF8 terminals.
    ws = wb.worksheets[2]
    rows = []
    for index, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None and text(value) for value in values):
            continue
        if len(values) < 9:
            raise ValueError(f"row {index}: expected 9 audit columns")
        name, matched_name, size, cny_price, usd_price, top_tab, category, direction, compatible = values[:9]
        product_name = text(name)
        if not product_name:
            raise ValueError(f"row {index}: 商品名称_文件名 is blank")
        rows.append({
            "row_number": index,
            "product_name": product_name,
            "matched_name": text(matched_name) or product_name,
            "size_mm": number(size, "size_mm", index),
            "usd_price": computed_usd_price(usd_price, cny_price, index),
            "top_tab": text(top_tab),
            "category": text(category),
            "direction_rule": normalized_direction(text(direction) or "none"),
            "compatible_bead_sizes": text(compatible),
        })
    return rows


def component_fields(top_tab: str) -> tuple[str, str]:
    if top_tab in {"水晶", "天然石"}:
        return "crystal", "crystal"
    if top_tab in {"配饰", "随型"}:
        return "accessory", "accessory"
    raise ValueError(f"Unsupported first-level tab: {top_tab!r}")


def build_manifest(rows: list[dict[str, object]], images: dict[str, Path]) -> tuple[list[dict[str, str]], list[str]]:
    by_product: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        by_product[str(row["product_name"])].append(row)
    manifest = []
    missing = []
    for product_name in by_product:
        source_name = WORKBOOK_TO_SOURCE.get(product_name, product_name)
        image = images.get(product_name) or images.get(source_name)
        if not image:
            missing.append(product_name)
            continue
        material_key = stable_key(product_name)
        manifest.append({
            "material_key": material_key,
            "source_product_name_zh": product_name,
            "source_image_path": str(image),
            "upload_filename": f"{material_key}.webp",
            "wordpress_attachment_id": "",
            "wordpress_media_url": "",
            "upload_status": "pending",
        })
    return manifest, missing


def write_manifest(path: Path, manifest: list[dict[str, str]]) -> None:
    headers = list(manifest[0].keys()) if manifest else [
        "material_key", "source_product_name_zh", "source_image_path", "upload_filename",
        "wordpress_attachment_id", "wordpress_media_url", "upload_status",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(manifest)


def read_manifest(path: Path) -> dict[str, dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return {row["source_product_name_zh"]: row for row in csv.DictReader(handle)}


def build_catalog(rows: list[dict[str, object]], manifest: dict[str, dict[str, str]]) -> tuple[list[dict[str, str]], list[str]]:
    crystal_sizes = sorted({float(row["size_mm"]) for row in rows if component_fields(str(row["top_tab"]))[0] == "crystal"})
    material_order: dict[str, int] = {}
    output: list[dict[str, str]] = []
    missing_media: list[str] = []
    seen_variants: set[str] = set()
    for row in rows:
        name = str(row["product_name"])
        media = manifest.get(name)
        if not media or not text(media.get("wordpress_media_url")):
            missing_media.append(name)
            continue
        component_type, tab_slug = component_fields(str(row["top_tab"]))
        if name not in material_order:
            material_order[name] = len(material_order) + 1
        material_key = stable_key(name)
        size = float(row["size_mm"])
        variant_key = f"{material_key}-{quantity(size)}mm"
        if variant_key in seen_variants:
            raise ValueError(f"Duplicate product/size Variant key: {variant_key}")
        seen_variants.add(variant_key)
        direction = str(row["direction_rule"])
        shape = "round" if component_type == "crystal" else ("charm" if direction.startswith("radial") else "accessory")
        output.append({
            "material_key": material_key,
            "component_type": component_type,
            "library_tab_slug": tab_slug,
            "category_slug": category_key(str(row["category"])),
            "name_en": name,
            "primary_color": "",
            "color_tags": "",
            "intention_tags": "",
            "material_image_url": text(media["wordpress_media_url"]),
            "material_status": "live",
            "material_sort_order": str(material_order[name]),
            "variant_key": variant_key,
            "size_mm": quantity(size),
            "shape": shape,
            "price": quantity(float(row["usd_price"])),
            "occupied_length_mm": quantity(size),
            "display_scale": "1",
            "variant_image_url": text(media["wordpress_media_url"]),
            "stock_status": "onbackorder",
            "stock_quantity": "",
            "compatibility": "",
            "compatible_bead_sizes": parse_compatibility(str(row["compatible_bead_sizes"]), crystal_sizes) if component_type == "accessory" else "",
            "orientation_mode": direction,
            "mirrored_variant_key": "",
            "allowed_orientations": "",
            "allowed_positions": "",
            "neighbor_constraints": "",
            "variant_status": "live",
            "variant_sort_order": str(row["row_number"]),
            "source_name": "linganshi-reference-20260722",
        })
    return output, sorted(set(missing_media))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--media-dir", type=Path, default=DEFAULT_MEDIA_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--build-catalog", action="store_true", help="Require uploaded media URLs and write catalog-import.json")
    parser.add_argument("--allow-missing-media", action="store_true", help="With --build-catalog, write only rows that have uploaded media and report withheld products")
    args = parser.parse_args()

    if not args.workbook.is_file():
        raise SystemExit(f"Workbook not found: {args.workbook}")
    if not args.media_dir.is_dir():
        raise SystemExit(f"Media directory not found: {args.media_dir}")
    args.output_dir.mkdir(parents=True, exist_ok=True)
    rows = load_rows(args.workbook)
    images = image_stems(args.media_dir)
    manifest, missing_images = build_manifest(rows, images)
    manifest_path = args.output_dir / "linganshi-media-manifest.csv"
    if not manifest_path.exists():
        write_manifest(manifest_path, manifest)
    else:
        # Preserve successful media URLs/statuses across reruns while refreshing
        # source paths and rows from the reviewed workbook.
        previous = read_manifest(manifest_path)
        for item in manifest:
            item.update({key: previous.get(item["source_product_name_zh"], {}).get(key, item[key]) for key in ("wordpress_attachment_id", "wordpress_media_url", "upload_status")})
        write_manifest(manifest_path, manifest)

    print(f"reviewed_variant_rows={len(rows)}")
    print(f"unique_webp_images={len(images)}")
    print(f"media_manifest_products={len(manifest)}")
    print(f"products_without_local_webp={len(missing_images)}")
    for item in missing_images:
        print(f"MISSING_LOCAL_WEBP\t{item}")
    if not args.build_catalog:
        print(f"manifest={manifest_path}")
        return 0

    catalog, missing_media = build_catalog(rows, read_manifest(manifest_path))
    if missing_media:
        print(f"catalog_withheld_products_missing_wordpress_media={len(missing_media)}", file=sys.stderr)
        for item in missing_media:
            print(f"MISSING_WORDPRESS_MEDIA\t{item}", file=sys.stderr)
        if not args.allow_missing_media:
            return 2
    payload_path = args.output_dir / "linganshi-catalog-import.json"
    payload_path.write_text(json.dumps({"rows": catalog}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"catalog_variant_rows={len(catalog)}")
    print(f"catalog_payload={payload_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
